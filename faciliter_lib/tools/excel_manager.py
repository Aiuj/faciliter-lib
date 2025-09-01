"""
Module for loading and manipulating Excel files, and rendering their contents as markdown tables.
"""

from openpyxl import load_workbook
from tabulate import tabulate
from faciliter_lib import get_module_logger
import os

# Get module-specific logger
logger = get_module_logger()

class ExcelManager:
    """
    Manages loading, rendering, and manipulation of Excel workbooks.
    """

    def __init__(self, excel_path):
        """
        Initialize the ExcelManager with the path to the Excel file.

        Args:
            excel_path (str): Path to the Excel file.
        """
        self.excel_path = excel_path
        self.wb = None
        logger.debug(f"ExcelManager initialized with path: {excel_path}")

    def load(self):
        """
        Loads the Excel workbook from the specified path.

        Returns:
            Workbook: The loaded openpyxl Workbook object.
        """
        try:
            logger.info(f"Loading Excel workbook from: {self.excel_path}")
            self.wb = load_workbook(self.excel_path, read_only=True)
            logger.info(f"Excel workbook loaded successfully - sheets: {self.wb.sheetnames}")
            return self.wb
        except Exception as e:
            logger.error(f"Failed to load Excel workbook from {self.excel_path}: {str(e)}")
            raise

    # Replace None and NaN with empty string
    def clean_cell(self, cell):
        if cell is None:
            return ''
        try:
            if isinstance(cell, float) and str(cell) == 'nan':
                return ''
        except Exception:
            pass
        return cell

    def get_sheet_tables(self, ws, max_rows=None, add_col_headers=False, add_row_headers=False):
        """
        Extracts sheet data as a list of lists, with options to add Excel-style column and row headers.
        Preserves original cell locations for headers, even if empty rows/columns are removed.

        Args:
            ws: Worksheet object.
            max_rows (int, optional): Maximum number of rows to return.
            add_col_headers (bool): If True, adds Excel-style column headers (A, B, ...).
            add_row_headers (bool): If True, adds Excel-style row headers (1, 2, ...).

        Returns:
            list: Sheet data as a list of lists, with optional headers.
        """
        # Get all rows as lists, preserving original shape
        data = list(ws.values)
        num_rows = len(data)
        num_cols = max((len(row) for row in data), default=0)
        # Normalize all rows to same length
        norm_data = [list(row) + [''] * (num_cols - len(row)) for row in data]

        # Track which rows/cols are empty
        empty_row_flags = [not any(cell not in (None, '', float('nan')) for cell in row) for row in norm_data]
        empty_col_flags = [not any(norm_data[row_idx][col_idx] not in (None, '', float('nan')) for row_idx in range(num_rows)) for col_idx in range(num_cols)]

        # Remove empty rows/cols but keep original indices for headers
        row_map = [i for i, empty in enumerate(empty_row_flags) if not empty]
        col_map = [i for i, empty in enumerate(empty_col_flags) if not empty]
        filtered_data = [[norm_data[row_idx][col_idx] for col_idx in col_map] for row_idx in row_map]

        # Limit rows
        if max_rows is not None:
            filtered_data = filtered_data[:max_rows]
            row_map = row_map[:max_rows]

        # Clean cells
        filtered_data = [[self.clean_cell(cell) for cell in row] for row in filtered_data]

        # Add column headers (A, B, C, ...)
        if add_col_headers and filtered_data:
            def excel_col_name(n):
                name = ''
                while n > 0:
                    n, r = divmod(n-1, 26)
                    name = chr(65 + r) + name
                return name
            col_headers = [excel_col_name(i+1) for i in col_map]
            if add_row_headers:
                col_headers = [''] + col_headers
            filtered_data.insert(0, col_headers)

        # Add row headers (1, 2, ...)
        if add_row_headers and filtered_data:
            for idx, row in enumerate(filtered_data):
                # If col headers are present, skip the first row
                if add_col_headers and idx == 0:
                    continue
                # Use original Excel row number (1-based)
                row_num = str(row_map[idx-1]+1) if add_col_headers else str(row_map[idx]+1)
                row.insert(0, row_num)

        return filtered_data
    
    def get_content(self, max_rows=None, add_col_headers=True, add_row_headers=True, detect_language=True):
        """
        Returns the content of the workbook as a markdown string.

        Args:
            max_rows (int, optional): Maximum number of rows to render per sheet.
            add_col_headers (bool): If True, adds Excel-style column headers (A, B, ...).
            add_row_headers (bool): If True, adds Excel-style row headers (1, 2, ...).
            detect_language (bool): If True, detects language of the content.

        Returns:
            list: List of dictionaries containing sheet name, markdown content, language, and rows.
        
        Raises:
            ValueError: If the workbook is not loaded.
        """
        from faciliter_lib.utils.language_utils import LanguageUtils

        if self.wb is None:
            logger.error("Workbook not loaded. Cannot get content.")
            raise ValueError("Workbook not loaded. Call load() first.")

        logger.info(f"Getting content from workbook - max_rows: {max_rows}")
        results = []
        
        for sheet_name in self.wb.sheetnames:
            logger.debug(f"Processing sheet: {sheet_name}")
            ws = self.wb[sheet_name]
            data = self.get_sheet_tables(ws, max_rows, add_col_headers=add_col_headers, add_row_headers=add_row_headers)
            # Prepare headers
            headers = data[0] if data else []
            rows = data[1:] if len(data) > 1 else []
            md_table = tabulate(rows, headers=headers, tablefmt='github')
            
            if detect_language:
                # Detect language using faciliter-lib
                language = LanguageUtils.detect_language(sheet_name + ":\n" + md_table)
                logger.debug(f"Language detected for sheet '{sheet_name}': {language}")
            else:
                language = None
                
            results.append({
                "sheet_name": sheet_name,
                "markdown": md_table,
                "language": language,
                "rows": rows,
            })

        logger.info(f"Content processing completed for {len(results)} sheets")
        return results


    def to_markdown(self, max_rows=None, add_col_headers=True, add_row_headers=True, detect_language=True):
        """
        Converts all sheets in the loaded workbook to a list of markdown tables.

        Args:
            max_rows (int, optional): Maximum number of rows to render per sheet. Renders all rows if None.
            add_col_headers (bool): If True, adds Excel-style column headers (A, B, ...).
            add_row_headers (bool): If True, adds Excel-style row headers (1, 2, ...).
            detect_language (bool): If True, detects language of the content.

        Returns:
            list: List of dictionaries with keys:
                - 'sheet_name': Name of the sheet/tab
                - 'markdown': Markdown-formatted table for the sheet
                - 'language': Detected language (if detect_language=True)
        
        Raises:
            ValueError: If the workbook is not loaded.
        """
        from faciliter_lib.utils.language_utils import LanguageUtils

        if self.wb is None:
            logger.error("Workbook not loaded. Cannot convert to markdown.")
            raise ValueError("Workbook not loaded. Call load() first.")
            
        logger.info(f"Converting workbook to markdown - max_rows: {max_rows}")
        results = []
        
        for sheet_name in self.wb.sheetnames:
            logger.debug(f"Converting sheet to markdown: {sheet_name}")
            ws = self.wb[sheet_name]
            data = self.get_sheet_tables(ws, max_rows, add_col_headers=add_col_headers, add_row_headers=add_row_headers)
            # Prepare headers
            headers = data[0] if data else []
            rows = data[1:] if len(data) > 1 else []
            md_table = tabulate(rows, headers=headers, tablefmt='github')
            
            # Detect language if requested
            language = None
            if detect_language:
                try:
                    language = LanguageUtils.detect_language(sheet_name + ":\n" + md_table)
                    logger.debug(f"Language detected for sheet '{sheet_name}': {language}")
                except Exception as e:
                    logger.warning(f"Language detection failed for sheet '{sheet_name}': {e}")
                    language = None
            
            results.append({
                "sheet_name": sheet_name,
                "markdown": md_table,
                "language": language,
                "row_count": len(rows)
            })
            
        logger.info(f"Markdown conversion completed for {len(self.wb.sheetnames)} sheets")
        return results

    def to_combined_markdown(self, max_rows=None, add_col_headers=True, add_row_headers=True, detect_language=True):
        """
        Converts all sheets in the loaded workbook to a single combined markdown string with titles per tab.

        Args:
            max_rows (int, optional): Maximum number of rows to render per sheet. Renders all rows if None.
            add_col_headers (bool): If True, adds Excel-style column headers (A, B, ...).
            add_row_headers (bool): If True, adds Excel-style row headers (1, 2, ...).
            detect_language (bool): If True, detects language of the content.

        Returns:
            str: Combined markdown-formatted string with all sheets, each with its own title.
        
        Raises:
            ValueError: If the workbook is not loaded.
        """
        logger.info("Converting workbook to combined markdown")
        
        # Get individual sheet markdowns
        sheet_markdowns = self.to_markdown(
            max_rows=max_rows,
            add_col_headers=add_col_headers,
            add_row_headers=add_row_headers,
            detect_language=detect_language
        )
        
        # Combine them with titles
        combined_md = ''
        for sheet_data in sheet_markdowns:
            sheet_name = sheet_data['sheet_name']
            markdown = sheet_data['markdown']
            
            combined_md += f'## {sheet_name}\n\n'
            combined_md += markdown + '\n\n'
        
        logger.info(f"Combined markdown created for {len(sheet_markdowns)} sheets")
        return combined_md

    def to_json_ir(self, filename: str = None, max_rows: int = None) -> dict:
        """
        Build a structured JSON Intermediate Representation (IR) for the loaded workbook.

        Args:
            filename (str, optional): Source filename used for id/metadata.
            max_rows (int, optional): Maximum number of data rows (excluding header) to include per sheet.

        Returns:
            dict: { 'document': { ... }, 'language': detected_language }
        """
        from openpyxl.utils import range_boundaries, get_column_letter
        from tabulate import tabulate
        import uuid
        from faciliter_lib.utils.language_utils import LanguageUtils

        if self.wb is None:
            logger.error("Workbook not loaded. Cannot build JSON IR.")
            raise ValueError("Workbook not loaded. Call load() first.")

        base_name = os.path.splitext(os.path.basename(filename or self.excel_path))[0]
        short_id = uuid.uuid4().hex[:8]

        document = {
            "id": f"{base_name}_{short_id}",
            "type": "excel",
            "source_filename": os.path.basename(filename or self.excel_path),
            "language": None,
            "sheets": []
        }

        all_md_snippets = []

        full_wb = None

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            try:
                dim = ws.calculate_dimension()
                min_col, min_row, max_col, max_row = range_boundaries(dim)
            except Exception:
                dim = None
                min_col = min_row = max_col = max_row = None

            def _scan_bounds():
                nonlocal min_col, min_row, max_col, max_row
                min_col = None
                min_row = None
                max_col = None
                max_row = None
                try:
                    max_r = ws.max_row or 1
                    max_c = ws.max_column or 1
                    for r in range(1, max_r + 1):
                        row_has = False
                        row_min_c = None
                        row_max_c = None
                        for c in range(1, max_c + 1):
                            val = self.clean_cell(ws.cell(row=r, column=c).value)
                            if val not in (None, ''):
                                row_has = True
                                row_min_c = c if row_min_c is None else min(row_min_c, c)
                                row_max_c = c if row_max_c is None else max(row_max_c, c)
                        if row_has:
                            if min_row is None:
                                min_row = r
                            max_row = r
                            if row_min_c is not None:
                                min_col = row_min_c if min_col is None else min(min_col, row_min_c)
                            if row_max_c is not None:
                                max_col = row_max_c if max_col is None else max(max_col, row_max_c)
                except Exception:
                    pass

            if not (isinstance(min_col, int) and isinstance(min_row, int) and isinstance(max_col, int) and isinstance(max_row, int)) or (min_col == max_col == 1 and min_row == max_row == 1):
                _scan_bounds()

            if not all(isinstance(v, int) and v >= 1 for v in [min_col, min_row, max_col, max_row]) or (min_col == max_col == 1 and min_row == max_row == 1):
                try:
                    if full_wb is None:
                        from openpyxl import load_workbook as _lb
                        full_wb = _lb(self.excel_path, read_only=False, data_only=True)
                    full_ws = full_wb[sheet_name]
                    f_min_r = f_min_c = None
                    f_max_r = f_max_c = None
                    for row in full_ws.iter_rows(values_only=False):
                        row_has = False
                        r_index = None
                        r_min_c = None
                        r_max_c = None
                        for cell in row:
                            r_index = cell.row
                            val = self.clean_cell(cell.value)
                            if val not in (None, ''):
                                row_has = True
                                c_index = cell.column
                                r_min_c = c_index if r_min_c is None else min(r_min_c, c_index)
                                r_max_c = c_index if r_max_c is None else max(r_max_c, c_index)
                        if row_has:
                            if f_min_r is None:
                                f_min_r = r_index
                            f_max_r = r_index
                            if r_min_c is not None:
                                f_min_c = r_min_c if f_min_c is None else min(f_min_c, r_min_c)
                            if r_max_c is not None:
                                f_max_c = r_max_c if f_max_c is None else max(f_max_c, r_max_c)
                    if all(isinstance(v, int) and v >= 1 for v in [f_min_c, f_min_r, f_max_c, f_max_r]):
                        min_col, min_row, max_col, max_row = f_min_c, f_min_r, f_max_c, f_max_r
                except Exception:
                    pass

            if all(isinstance(v, int) and v >= 1 for v in [min_col, min_row, max_col, max_row]):
                dim = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            else:
                min_col = min_row = max_col = max_row = 1
                dim = 'A1:A1'

            columns = {}
            headers_for_md = []
            for col_idx in range(min_col, max_col + 1):
                col_letter = get_column_letter(col_idx)
                _ws_for_header = None
                try:
                    _ws_for_header = full_wb[sheet_name] if full_wb is not None else ws
                except Exception:
                    _ws_for_header = ws
                raw_header = _ws_for_header.cell(row=min_row, column=col_idx).value
                header_val = self.clean_cell(raw_header)
                header_val = header_val if (isinstance(header_val, str) and header_val.strip() != '') else (header_val if header_val not in (None, '') else f"Column {col_letter}")
                columns[col_letter] = {"header": header_val}
                headers_for_md.append(header_val)

            data_rows = []
            md_rows = []
            max_data_rows = (max_rows if max_rows is not None else (max_row - min_row))
            current_count = 0
            for row_idx in range(min_row + 1, max_row + 1):
                cells_obj = {}
                row_vals_md = []
                is_non_empty = False
                for col_idx in range(min_col, max_col + 1):
                    col_letter = get_column_letter(col_idx)
                    _ws_for_cell = None
                    try:
                        _ws_for_cell = full_wb[sheet_name] if full_wb is not None else ws
                    except Exception:
                        _ws_for_cell = ws
                    val = self.clean_cell(_ws_for_cell.cell(row=row_idx, column=col_idx).value)
                    if val not in (None, ''):
                        is_non_empty = True
                    cells_obj[col_letter] = val
                    row_vals_md.append(val)

                if not is_non_empty:
                    continue

                data_rows.append({
                    "row": row_idx,
                    "cells": cells_obj
                })
                md_rows.append(row_vals_md)
                current_count += 1
                if current_count >= max_data_rows:
                    break

            md_table = tabulate(md_rows, headers=headers_for_md, tablefmt='github') if headers_for_md else ''
            all_md_snippets.append(f"## {sheet_name}\n{md_table}")

            sheet_lang = None
            if md_table:
                try:
                    sheet_lang = LanguageUtils.detect_language(sheet_name + "\n" + md_table)
                except Exception:
                    sheet_lang = None

            block = {
                "block_id": f"b{len(document['sheets']) + 1}",
                "type": "table",
                "range": dim,
                "header_row": min_row,
                "columns": columns,
                "rows": data_rows,
                "text_md": md_table,
                "lang": sheet_lang,
            }

            document["sheets"].append({
                "name": sheet_name,
                "blocks": [block]
            })

        combined_md = "\n\n".join(all_md_snippets)
        try:
            overall_lang = LanguageUtils.detect_language(combined_md) if combined_md.strip() else None
        except Exception:
            overall_lang = None
        document["language"] = overall_lang

        result = {
            "document": document,
            "language": overall_lang,
        }

        logger.info("Workbook JSON IR built successfully")
        return result

